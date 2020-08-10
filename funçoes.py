#1- função de validação de dados numericos

def lernum(significado, tipo_valor, maximo, minimo, passo):
    valor = input("Introduza {0}: ".format(significado))
    while valor == '':
        print('Valor inválido!')
        valor = input("Introduza {0}: ".format(significado))
    while valor == ' ':
        print('Valor inválido!')
        valor = input("Introduza {0}: ".format(significado))
    if tipo_valor== 'f':
         if "," in valor:
            valor = valor.replace(",", ".")
         valor = float(valor)
         if minimo > maximo:
             print('Erro Tecnico! Parametros valor maximo ou minimo invalidos. Consultar o autor do programa')
         while valor > maximo:
             print('Valor invalido')
             valor = float(input("Introduza {0}: ".format(significado)))
         while valor < minimo:
             print('Valor invalido')
             valor = float( input("Introduza {0}: ".format(significado)))
         while passo > (maximo-minimo):
             print('Erro Tecnico! Parametro passso invalido. Consultar o autor do programa')
    elif tipo_valor == 'in':
        if "," in valor:
            print("Valor inválido. Introduza um valor inteiro")
        elif "." in valor:
            print("Valor inválido.Introduza um valor inteiro")
        valor = int(valor)
        if minimo > maximo:
             print('Erro Tecnico! Parametros valor maximo ou minimo invalidos. Consultar o autor do programa')
        while valor > maximo:
            print('Valor invalido')
            valor =int( input("Introduza {0}: ".format(significado)))
        while valor < minimo:
            print('Valor invalido')
            valor =int( input("Introduza {0}: ".format(significado)))
        while passo > (maximo-minimo):
            print('Erro Tecnico! Parametro passso invalido. Consultar o autor do programa')
    else:
        print('Tipo de valor nao reconhecido')
    return valor



#2- função de abreviiar nomes

def abreviar(nome_completo='?'):
    # descrição:
    """
       Abrevia um nome, da seguinte forma:
           - O primeiro e o último nomes ficam inalterados.
           - As partículas {'de', 'da', 'do', 'das', 'dos', 'e'} são removidas.
           - Os nomes intermédios são reduzidos à letra inicial em maiúscula,
             seguida dum ponto.
    """
    # VALIDAÇÃO TÉCNICA:
    if nome_completo == '?' or nome_completo.strip() == '':
        print('\nERRO TÉCNICO NA FUNÇÃO abrevia()!')
        print('    Argumento inválido ou não fornecido.')
        print('    Por favor contacte o responsável pelo programa.')
        return None
        
    # INICIALIZAÇÕES:
    abreviado = ''
    particulas = {'de', 'da', 'do', 'das', 'dos', 'e'}
    nomes = nome_completo.split()
    i = 0
    
    # TRANSFORMAÇÃO:
    for nome in nomes:            
        i = i+1            
        # PRIMEIRO NOME
        if i == 1:
            abreviado = nome.capitalize()
            continue        
        # NOMES INTERMÉDIOS
        elif i < len(nomes):
            if nome.lower() not in particulas:
                abreviado = abreviado + ' ' + nome[0].upper() + '.'
            continue            
        # ÚLTIMO NOME
        else:
            abreviado = abreviado + ' ' + nome.capitalize()

    return abreviado 



#3- função de validação de nomes

def lernome(pedido='?', formato='?', tmin=2, tmax=80):
    # descrição:
    """
       Pede um nome ao utilizador, especificando o significado do mesmo,
       valida-o e retorna-o num de 3 formatos possíveis:
           'n' - normal - primeira letra de cada nome em maiúscula
           'm' - maiúsculas - tudo em maiúsculas
           'f' - frase - apenas a primeira letra da primeira palavra em maiúscula
    """
    # VALIDAÇÃO TÉCNICA:
    if pedido == '?' or pedido.strip() == '' or formato == '?'or formato.strip() == '':
        print('\nERRO TÉCNICO NA FUNÇÃO lernome_v4()!')
        print('    Argumentos insuficientes.')
        print('    Por favor contacte o responsável pelo programa.')
        return None
    if formato not in {'n', 'm', 'f'}:
        print('\nERRO TÉCNICO NA FUNÇÃO lernome_v4()!')
        print('    Argumento formato="{0}" inválido.'.format(formato))
        print('    Por favor contacte o responsável pelo programa.')
        return None
    if tmin < 2 or tmax <= tmin:
        print('\nERRO TÉCNICO NA FUNÇÃO lernome_v4()!')
        print('    Argumentos de tamanho inválidos.')
        print('    Por favor contacte o responsável pelo programa.')
        return None
        
    # PEDIR O NOME AO UTILIZADOR E VALIDAR
    final = ''
    particulas = {'de', 'da', 'do', 'das', 'dos', 'e'}
    nome_ok = False
    while not nome_ok:
        introduzido = input('{0}: '.format(pedido).capitalize())
        nomes = introduzido.split()
        if len(nomes) == 0:
            print('\n"{0}" não pode ser vazio.'.format(pedido).capitalize())
            break
        i = 0
        for nome in nomes:            
            i = i+1            
            if not nome.isalpha() and formato != 'f':
                print('\nERRO: "{0}" só pode conter letras e espaços.'.format(pedido).capitalize())
                break
            # PRIMEIRO NOME
            if i == 1 and i < len(nomes):
                if nome.lower() in particulas and formato != 'f':
                    print('\nERRO: O primeiro nome não pode ser "{0}".'.format(nome))
                    break                   
                elif formato == 'm':
                    final = nome.upper()
                    continue
                else:
                    final = nome.capitalize()
                    continue            
            # NOMES INTERMÉDIOS
            elif i < len(nomes):
                if formato == 'm':
                    final = final + ' ' + nome.upper()
                    continue
                elif nome.lower() in particulas or formato == 'f':
                    final = final + ' ' + nome.lower()
                    continue                   
                else:
                    final = final + ' ' + nome.capitalize()
                    continue            
            # ÚLTIMO NOME
            else:
                if nome.lower() in particulas:
                    print('\nERRO: O último nome não pode ser "{0}".'.format(nome))                
                elif formato == 'm':
                    final = final + ' ' + nome.upper()
                    nome_ok = True
                elif formato == 'f':
                    final = final + ' ' + nome.lower()
                    nome_ok = True
                else:
                    final = final + ' ' + nome.capitalize()
                    nome_ok = True
                    
        # VALIDAR TAMANHO E, SE NECESSÁRIO, SUGERIR ABREVIATURA
        if nome_ok == True:
            if len(final) < tmin:
                print('\nERRO: O tamanho mínimo para "{0}" é de {1} carateres.'.format(pedido, tmin).capitalize())
                nome_ok = False
            elif len(final) > tmax:
                print('\nERRO: O tamanho máximo para "{0}" é de {1} carateres.'.format(pedido, tmax).capitalize())
                nome_ok = False
                final = abreviar(final)
                if formato in {'n', 'm'} and len(final) <= tmax:
                    resposta = input('Aceita abreviar para "{0}" (sim/não)? '.format(final))
                    if resposta[0].lower() == 's':
                        nome_ok = True
                    else:
                        final = ''
        
    return final



#4 - função de validação de números mecanográficos 

def num_mec(tipo):
    mec = input('Introduza o número mecanográfico do\da {0}: '.format(tipo))
    if tipo == 'Condutor':
        while not 'Cd' in mec:
            print('O número mecanográfico não corresponde a {0}!'.format(tipo))
            mec = input('Introduza número mecanográfico do\da {0}: '.format(tipo))
    elif tipo == 'Viagem':
        while not 'V' in mec:
            print('O número mecanográfico não corresponde a {0}!'.format(tipo))
            mec = input('Introduza número mecanográfico do\da {0}: '.format(tipo))
    elif tipo == 'Contentor':
        while not 'Ct' in mec:
            print('O número mecanográfico não corresponde a {0}!'.format(tipo))
            mec = input('Introduza número mecanográfico do\da {0}: '.format(tipo))
    elif tipo == 'Rota':
        while not 'R' in mec:
            print('O número mecanográfico não corresponde a {0}!'.format(tipo))
            mec = input('Introduza número mecanográfico do\da {0}: '.format(tipo))
    else:
        print('Erro técnico! Argumento formato="{0}" inválido.'.format(tipo))
    
    return mec



#5: função eliminar dados

def eliminar(worksheet,significado):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb.get_sheet_by_name(worksheet)
    continuar = 1
    while continuar != 0:
        if worksheet == 'Condutores':
            num = num_mec(significado)
            begrow = 1
            endrow = ws.max_row
            existe = False
            for i in range (begrow,endrow):
                if ws.cell(row=i, column=2).value == num:
                    existe = True
                    ws.cell(row=i, column=1).value = 0
                    ws.cell(row=i, column=2).value = 0
                    ws.cell(row=i, column=3).value = 0
                    ws.cell(row=i, column=4).value = 0
                    ws.cell(row=i, column=5).value = 0
                    ws.cell(row=i, column=6).value = 0
                    ws.cell(row=i, column=7).value = 0
                    ws.cell(row=i, column=8).value = 0
                    ws.row_dimensions[i].hidden = True
                    break
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado))      
        elif worksheet == 'Viagens':
            num = num_mec(significado)
            begrow = 1
            endrow = ws.max_row
            existe = False
            for i in range (begrow,endrow):
                if ws.cell(row=i, column=1).value == num:
                    existe = True
                    ws.cell(row=i, column=1).value = 0
                    ws.cell(row=i, column=2).value = 0
                    ws.cell(row=i, column=3).value = 0
                    ws.cell(row=i, column=4).value = 0
                    ws.cell(row=i, column=5).value = 0
                    ws.cell(row=i, column=6).value = 0
                    ws.cell(row=i, column=7).value = 0
                    ws.cell(row=i, column=8).value = 0
                    ws.cell(row=i, column=9).value = 0
                    ws.cell(row=i, column=10).value = 0
                    ws.cell(row=i, column=11).value = 0
                    ws.cell(row=i, column=12).value = 0
                    ws.row_dimensions[i].hidden = True
                    break
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado)) 
                                     
        elif worksheet == 'Carrinhas': 
            num = input('Introduza a matrícula da carrinha: ')
            begrow = 1
            endrow = ws.max_row
            existe = False
            for i in range (begrow,endrow):
                if ws.cell(row=i, column=1).value == num:
                    existe = True
                    ws.cell(row=i, column=1).value = 0
                    ws.cell(row=i, column=2).value = 0
                    ws.cell(row=i, column=3).value = 0
                    ws.cell(row=i, column=4).value = 0
                    ws.row_dimensions[i].hidden = True
                    break
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado))  
                                     
        elif worksheet == 'Contentores':
            num = num_mec(significado)
            begrow = 1
            endrow = ws.max_row
            existe = False
            for i in range (begrow,endrow):
                if ws.cell(row=i, column=1).value == num:
                    existe = True
                    ws.cell(row=i, column=1).value = 0
                    ws.cell(row=i, column=2).value = 0
                    ws.cell(row=i, column=3).value = 0
                    ws.cell(row=i, column=4).value = 0
                    ws.cell(row=i, column=5).value = 0
                    ws.row_dimensions[i].hidden = True
                    break
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado)) 
                                     
        elif worksheet == 'Rotas':
            num = num_mec(significado)
            begrow = 1
            endrow = ws.max_row
            existe = False
            for i in range (begrow,endrow):
                if ws.cell(row=i, column=1).value == num:
                    existe = True
                    ws.cell(row=i, column=1).value = 0
                    ws.cell(row=i, column=2).value = 0
                    ws.cell(row=i, column=3).value = 0
                    ws.cell(row=i, column=4).value = 0
                    ws.cell(row=i, column=5).value = 0
                    ws.row_dimensions[i].hidden = True
                    break
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado)) 
                                     
        print('{:*^50}'.format(''))
        print('Deseja eliminar mais {0}? '.format(worksheet))
        print('1: sim      0:nao')
        print('{:*^50}'.format(''))
        continuar = int(input('Opção: '))
    wb.save('Gestao de transportes.xlsx')
    
    

#6: função adicionar dados

def adicionar(worksheet):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb.get_sheet_by_name(worksheet)
    continuar = 1
    while continuar ==  1:
        b = ws.max_row # última linha da worksheet
        a = b + 1
        if worksheet == 'Condutores':
            ws.cell(row=a, column=1).value = lernome('nome do condutor','n',2,80)
            ws.cell(row=a, column=2).value = num_mec('Condutor')
            ws.cell(row=a, column=3).value = lernum('idade do condutor','in',200,0,1)
            ws.cell(row=a, column=4).value = lernum('número de atrasos do condutor','in',100,0,1)
            ws.cell(row=a, column=5).value = lernum('número de faltas do condutor','in',100,0,1)
            ws.cell(row=a, column=6).value = lernum('número de quilómetros percorridos pelo condutor','f',100000000000000000,0,1)
            ws.cell(row=a, column=7).value = lernum('número de acidentes do condutor','in',100,0,1)
            ws.cell(row=a, column=8).value = input('Introduza o período de férias do condutor: ')
            print('{:*^50}'.format(''))
            print('Deseja adicionar mais {0}? '.format(worksheet))
            print('1: sim      0:nao')
            print('{:*^50}'.format(''))
            continuar = int(input('Opção: '))
            break 
        
        if worksheet == 'Viagens':
            ws.cell(row=a, column=1).value = num_mec('Viagem')
            ws.cell(row=a, column=2).value = input('Introduza a data e hora de partida? ')
            ws.cell(row=a, column=3).value = input('Introduza a data e hora de chegada? ')
            ws.cell(row=a, column=4).value = lernome('origem','n',2,80)
            ws.cell(row=a, column=5).value = lernome('destino','n',2,80)
            ws.cell(row=a, column=6).value = num_mec('Condutor')
            ws.cell(row=a, column=7).value = input('Introduza o matrícula da carrinha? ')
            ws.cell(row=a, column=8).value = num_mec('Rota')
            ws.cell(row=a, column=9).value = lernum('número de ocorrências','in',100,0,1)
            ws.cell(row=a, column=10).value = lernome('a mercadoria','f',2,80)
            ws.cell(row=a, column=11).value = lernum('o número de pausas','in',100,0,1)
            ws.cell(row=a, column=12).value = lernum('o volume de mercadoria','f',10000000000000,0,0.0001)
            print('{:*^50}'.format(''))
            print('Deseja adicionar mais {0}? '.format(worksheet))
            print('1: sim      0:nao')
            print('{:*^50}'.format(''))
            continuar = int(input('Opção: '))
            break 
        
        if worksheet == 'Carrinhas':
            ws.cell(row=a, column=1).value = input('Introduza a matricula? ')
            ws.cell(row=a, column=2).value = lernum('número de km percorridos','f',100000000000000000000,0,0.0001)
            ws.cell(row=a, column=3).value = lernum('volume de mercadoria','f',10000000000000,0,0.0001)
            ws.cell(row=a, column=4).value = lernum('número de acidentes','in',100,0,1)
            print('{:*^50}'.format(''))
            print('Deseja adicionar mais {0}? '.format(worksheet))
            print('1: sim      0:nao')
            print('{:*^50}'.format(''))
            continuar = int(input('Opção: '))
            break 
        if worksheet == 'Contentores':
            ws.cell(row=a, column=1).value = num_mec('Contentor')
            ws.cell(row=a, column=2).value = lernum('a capacidade','f', 100000000000000000,0,0.0001)
            ws.cell(row=a, column=3).value = lernome('localização','n',2,80)
            ws.cell(row=a, column=4).value = num_mec('Rota')
            ws.cell(row=a, column=5).value = input('Introduza a data da última descarga: ')
            print('{:*^50}'.format(''))
            print('Deseja adicionar mais {0}? '.format(worksheet))
            print('1: sim      0:nao')
            print('{:*^50}'.format(''))
            continuar = int(input('Opção: '))
            break 
        
        if worksheet == 'Rotas':
            ws.cell(row=a, column=1).value = num_mec('Rota')
            ws.cell(row=a, column=2).value = lernome('local de início','n',2,80)
            ws.cell(row=a, column=3).value = lernome('local de fim','n',2,80)
            ws.cell(row=a, column=4).value = lernum('tempo médio necessário para percorrer a rota','f',24,0,0.0001)
            ws.cell(row=a, column=5).value = lernum('número km a percorrer','f',100000000000000,0,0.0001)
            print('{:*^50}'.format(''))
            print('Deseja adicionar mais {0}? '.format(worksheet))
            print('1: sim      0:nao')
            print('{:*^50}'.format(''))
            continuar = int(input('Opção: '))
            break 
        
        else:
            print('Erro técinico! Worksheet não atribuída. Consulte o responsável pelo programa.')
    wb.save('Gestao de transportes.xlsx')



#7: função alterar dados

def alterar(worksheet, significado):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb.get_sheet_by_name(worksheet)
    continuar = 1
    while continuar == 1:
        if worksheet == 'Condutores':
            num = num_mec(significado)
            a = ws.max_row
            existe = False
            for i in range(1,a):
                if ws.cell(row=i, column=2).value == num:
                    b=i
                    existe = True
                    terminar = False
                    while terminar == False:
                        print('{:*^50}'.format('MENU'))
                        print('1: Alterar nome')
                        print('2: Alterar número interno')
                        print('3: Alterar idade')
                        print('4: Alterar atrasos')
                        print('5: Alterar faltas')
                        print('6: Alterar quilómetros percorridos')
                        print('7: Alterar acidentes')
                        print('8: Alterar período de férias')
                        print('9: Terminar')
                        print('{:*^50}'.format(''))
                        op = int(input('Opção: '))
                        if op == 1:
                            ws.cell(row=b, column=1).value = lernome('nome do condutor','n',2,80)
                        elif op == 2:
                            ws.cell(row=b, column=2).value = num_mec('Condutor')
                        elif op == 3:
                            ws.cell(row=b, column=3).value = lernum('idade do condutor','in',150,0,1)
                        elif op == 4:
                            ws.cell(row=b, column=4).value = lernum('número de atrasos do condutor','in',100,0,1)
                        elif op == 5:
                            ws.cell(row=b, column=5).value = lernum('número de faltas do condutor','in',100,0,1)
                        elif op == 6: 
                            ws.cell(row=b, column=6).value = lernum('número de quilómetros percorridos pelo condutor','f',100000000000000000,0,1)
                        elif op == 7:
                            ws.cell(row=b, column=7).value = lernum('número de acidentes do condutor','in',100,0,1)
                        elif op == 8:
                            ws.cell(row=b, column=8).value = input('Introduza o período de férias do condutor? ')
                        elif op == 9:
                            terminar = True
                        else:
                            print('Opção inválida!')
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado))
        
        if worksheet == 'Viagens':
            num = num_mec(significado)
            a = ws.max_row
            existe = False
            for i in range(1,a):
                if ws.cell(row=i, column=1).value == num:
                    b=i
                    existe = True
                    terminar= False
                    while terminar == False:
                        print('{:*^50}'.format('MENU'))
                        print('1: Alterar número de identificação')
                        print('2: Alterar data e hora de partida')
                        print('3: Alterar data e hora de chegada')
                        print('4: Alterar origem')
                        print('5: Alterar destino')
                        print('6: Alterar número interno do condutor')
                        print('7: Alterar matrícula da carrinha')
                        print('8: Alterar número de identificação da rota')
                        print('9: Alterar ocorrências')
                        print('10: Alterar mercadoria')
                        print('11: Alterar pausas')
                        print('12: Alterar volume da mercadoria')
                        print('13: Terminar')
                        print('{:*^50}'.format(''))
                        op = int(input('Opção: '))
                        if op == 1:
                            ws.cell(row=b, column=1).value = num_mec('Viagem')
                        elif op == 2:
                            ws.cell(row=b, column=2).value = input('Introduza a data e hora de partida? ')
                        elif op == 3:
                            ws.cell(row=b, column=3).value = input('Introduza a data e hora de chegada? ')
                        elif op == 4:
                            ws.cell(row=b, column=4).value = lernome('origem','n',2,80)
                        elif op == 5:
                            ws.cell(row=b, column=5).value = lernome('destino','n',2,80)
                        elif op == 6:
                            ws.cell(row=b, column=6).value = num_mec('Condutor')
                        elif op == 7:
                            ws.cell(row=b, column=7).value = input('Introduza a matrícula da carrinha? ')
                        elif op == 8:
                            ws.cell(row=b, column=8).value = num_mec('Rota')
                        elif op == 9:
                            ws.cell(row=b, column=9).value = lernum('número de ocorrências','in',100,0,1)
                        elif op == 10:
                            ws.cell(row=b, column=10).value = lernome('a mercadoria','f',2,80)
                        elif op == 11:
                            ws.cell(row=b, column=11).value = lernum('o número de pausas','in',100,0,1)
                        elif op == 12:
                            ws.cell(row=b, column=12).value = lernum('o volume de mercadoria','f',10000000000,0,0.0001)
                        elif op == 13:
                            terminar = True
                        else:
                            print('Opção inválida!')
                    break
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado))
                
        if worksheet == 'Carrinhas':
            num = input('Introduza a matrícula da carrinha: ') 
            a = ws.max_row
            existe = False
            for i in range(1,a):
                if ws.cell(row=i, column=1).value == num:
                    b=i
                    existe = True
                    terminar= False
                    while terminar == False:
                        print('{:*^50}'.format('MENU'))
                        print('1: Alterar matrícula')
                        print('2: Alterar quilómetros percorridos')
                        print('3: Alterar volume da carga')
                        print('4: Alterar número de acidentes')
                        print('5: Terminar')
                        print('{:*^50}'.format(''))
                        op = int(input('Opção: '))
                        if op == 1:
                            ws.cell(row=b, column=1).value = input('Introduza o número matrícula: ')
                        elif op == 2:
                            ws.cell(row=b, column=2).value = lernum('número de quilómetros percorridos','f',100000000000,0,0.0001)
                        elif op == 3:
                            ws.cell(row=b, column=3).value = lernum('volume da carga em metros cúbicos','f',10000000000000,0,0.0001)
                        elif op == 4:
                            ws.cell(row=b, column=4).value = lernum('número de acidentes','in',100,0,1)
                        elif op == 5:
                            terminar == True
                        else:
                            print('Opção inválida!')
                    break
            if existe == False: 
                print('Número mecanográfico do\da {0} não atribído'.format(significado))
                
        if worksheet == 'Contentores':
            num = num_mec(significado)
            a = ws.max_row
            existe = False
            for i in range(1,a):
                if ws.cell(row=i, column=1).value == num:
                    b=i
                    existe = True
                    terminar= False
                    while terminar == False:
                        print('{:*^50}'.format('MENU'))
                        print('1: Alterar número de identificação do contentor')
                        print('2: Alterar capacidade')
                        print('3: Alterar localização')
                        print('4: Alterar número de identificação da rota')
                        print('5: Alterar data da última descarga')
                        print('6: Terminar')
                        print('{:*^50}'.format(''))
                        op = int(input('Opção: '))
                        if op == 1:
                            ws.cell(row=b, column=1).value = num_mec('Contentor')
                        elif op == 2:
                            ws.cell(row=b, column=2).value = lernum('capacidade em metros cúbicos','f','f', 100000000000000000,0,0.0001)
                        elif op == 3:
                            ws.cell(row=b, column=3).value = lernome('localização','n',2,80)
                        elif op == 4:
                            ws.cell(row=b, column=4).value = num_mec('Rota')
                        elif op == 5:
                            ws.cell(row=b, column=5).value = input('Introduza a data da última descarga: ')
                        elif op == 6:
                            terminar == True
                        else:
                            print('Opção inválida!')
                    break
            if existe == False: 
                print('Número mecanográfico do/da {0} não atribído'.format(significado))
                
        if worksheet == 'Rotas':
            num = num_mec(significado)
            a = ws.max_row
            existe = False
            for i in range(1,a):
                if ws.cell(row=i, column=1).value == num:
                    b=i
                    existe = True
                    terminar= False
                    while terminar == False:
                        print('{:*^50}'.format('MENU'))
                        print('1: Alterar número de identificação da rota')
                        print('2: Alterar local de início')
                        print('3: Alterar local de fim')
                        print('4: Alterar tempo médio necessário para percorrer a rota')
                        print('5: Alterar quilómetros a percorrer')
                        print('6: Terminar')
                        print('{:*^50}'.format(''))
                        op = int(input('Opção: '))
                        if op == 1:
                            ws.cell(row=b, column=1).value = num_mec('Rota')
                        elif op == 2:
                            ws.cell(row=b, column=2).value = lernome('local de início','n',2,80)
                        elif op == 3:
                            ws.cell(row=b, column=3).value = lernome('local de fim','n',2,80)
                        elif op == 4:
                            ws.cell(row=b, column=4).value = lernum('tempo médio necessário para percorrer a rota em horas','f',24,0,0.0001)
                        elif op == 5:
                            ws.cell(row=b, column=5).value = lernum('número km a percorrer','f',100000000000000,0,0.0001)
                        elif op == 6:
                            terminar == True
                        else:
                            print('Opção inválida!')
                    break
            if existe == False:
                print('Número mecanográfico do/da {0} não atribído'.format(significado))
                
        print('{:*^50}'.format(''))
        print('Deseja alterar mais {0}? '.format(worksheet))
        print('1: sim      0:nao')
        print('{:*^50}'.format(''))
        continuar = int(input('Opção: '))
    wb.save('Gestao de transportes.xlsx')



#8: função máximo

def maxi(categoria, parametro, identificacao, det1, det2, Col1, Col2):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb[categoria]
    
    coluna1 = ws[Col1]
    coluna2 = ws[Col2] 
    a = coluna1[1].value
    b = coluna2[1].value
    for i in range(2, len(coluna1)):
        if b < coluna2[i].value:
            a = coluna1[i].value
            b = coluna2[i].value
    print('{0} {1} {2} que tem mais {3} ({4}) tem como {5} {6}.'.format(det1, categoria, det2, parametro, b, identificacao, a))



#9: função mínimo 
   
def mini(categoria, parametro, identificacao, det1, det2, Col1, Col2):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb[categoria]
    
    coluna1 = ws[Col1]
    coluna2 = ws[Col2] 
    a = coluna1[1].value
    b = coluna2[1].value
    for i in range(2, len(coluna1)):
        if b > coluna2[i].value:
            a = coluna1[i].value
            b = coluna2[i].value
    print('{0} {1} {2} que tem menos {3} ({4}) tem como {5} {6}.'.format(det1, categoria, det2, parametro, b, identificacao, a))



#10: função ordenação

def ordenacao(categoria, col1, col2):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb[categoria]
    
    coluna1 = ws[col1]
    coluna2 = ws[col2]
    for i in range(1, len(coluna1)-1):
        for j in range(i+1, len(coluna1)):
            if coluna2[i].value > coluna2[j].value:
                coluna1[i].value, coluna1[j].value = coluna1[j].value, coluna1[i].value
                coluna2[i].value, coluna2[j].value = coluna2[j].value, coluna2[i].value
    for i in range(len(coluna1)):
        print(str(coluna1[i].value).center(20), str(coluna2[i].value).center(4))



#11: função procura

def procura(categoria, col):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb[categoria]

    coluna = ws[col]
    encontrada = False
    linha = ws['1']
    while not encontrada:
        pesquisa = input('Pesquisar: ')
        k = 0
        for i in range(1, len(coluna)):
            k = k+1
            if pesquisa == coluna[i].value:
                encontrada = True
                break
        if not encontrada:
            print('Não existem registos')
            
    for j in range(len(linha)):
        print(str(linha[j].value).center(40), str(ws[k+1][j].value).center(14))



#12: função mostrar informações

def info(categoria):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb[categoria]

    linha = ws['1']
    coluna = ws['A']
    colunas = len(linha)

    if colunas == 8:
        for i in range(1, len(coluna)+1):
            print(str(ws[i][0].value).center(20), str(ws[i][1].value).center(26), str(ws[i][2].value).center(7), str(ws[i][3].value).center(9), str(ws[i][4].value).center(8), str(ws[i][5].value).center(16), str(ws[i][6].value).center(9), str(ws[i][7].value).center(19))
    elif colunas == 12:
        for i in range(1, len(coluna)+1):
            print(str(ws[i][0].value).center(25), str(ws[i][1].value).center(26), str(ws[i][2].value).center(26), str(ws[i][3].value).center(8), str(ws[i][4].value).center(9), str(ws[i][5].value).center(26), str(ws[i][6].value).center(23), str(ws[i][7].value).center(21), str(ws[i][8].value).center(12), str(ws[i][9].value).center(12), str(ws[i][10].value).center(8), str(ws[i][11].value).center(20))
    elif colunas == 4:
        for i in range(1, len(coluna)+1):
            print(str(ws[i][0].value).center(11), str(ws[i][1].value).center(16), str(ws[i][2].value).center(17), str(ws[i][3].value).center(11))
    elif colunas == 5:
        for i in range(1, len(coluna)+1):
            print(str(ws[i][0].value).center(27), str(ws[i][1].value).center(17), str(ws[i][2].value).center(14), str(ws[i][3].value).center(29), str(ws[i][0].value).center(25))



#13: função média

def media(categoria, parametro, Col):

    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb[categoria]
    
    coluna = ws[Col] 
    soma = 0
    for i in range(1, len(coluna)):
        soma = soma + coluna[i].value
    media = soma/(len(coluna)-1)
    print('A média de {0} em relaçao a {1} é {2:.2f}.'.format(categoria, parametro, media))
    


#14: função para alterar a cor do separador de uma folha

def alterar_cor(worksheet):
    from openpyxl import load_workbook
    wb = load_workbook('Gestao de transportes.xlsx')
    ws = wb.get_sheet_by_name(worksheet)
    terminar = False
    while terminar == False:
        print('{:*^50}'.format('MENU'))
        print('1: Azul')
        print('2: Verde')
        print('3: Vermelho')
        print('4: Rosa')
        print('5: Amarelo')
        print('6: Cinza')
        print('7: Laranja')
        print('8: Terminar')
        print('{:*^50}'.format(''))
        op = int(input('Opção: '))
        if op == 1:
            ws.sheet_properties.tabColor = "1072BA"
        if op == 2:
            ws.sheet_properties.tabColor = "32CD32"
        if op == 3:
            ws.sheet_properties.tabColor = "DC143C"
        if op == 4:
            ws.sheet_properties.tabColor = "FF1493"
        if op == 5:
            ws.sheet_properties.tabColor = "FFD700"
        if op == 6:
            ws.sheet_properties.tabColor = "696969"
        if op == 7:
            ws.sheet_properties.tabColor = "FF4500"
        elif op == 8:
            terminar= True
            
    wb.save('Gestao de transportes.xlsx')